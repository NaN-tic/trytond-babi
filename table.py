import csv
import pytz
import time
import traceback
import datetime as mdatetime
from datetime import date, datetime, time as dt_time, timedelta
import logging
import sql
import unidecode
import json
import tempfile
import html
import urllib.parse
import secrets
from sql import Literal
from sql.operators import Equal
from decimal import Decimal
from types import SimpleNamespace
from openpyxl import Workbook
from openpyxl.writer.excel import save_workbook
from simpleeval import EvalWithCompoundTypes
from trytond import backend
from trytond.bus import notify
from trytond.transaction import Transaction
from trytond.pool import Pool
from trytond.model import (Exclude, Model, ModelView, ModelSQL, fields,
    Unique, DeactivableMixin, sequence_ordered, Workflow)
from trytond.exceptions import UserError
from trytond.i18n import gettext
from trytond.pyson import Bool, Eval, In, Not, PYSONDecoder, PYSONEncoder
from trytond.url import http_host
from trytond.config import config
from trytond.modules.company.model import (
    employee_field, reset_employee, set_employee)
from trytond.report import Report
from trytond.wizard import Wizard, StateView, StateAction, Button
from trytond.rpc import RPC
from .babi import TimeoutChecker, TimeoutException, FIELD_TYPES, QUEUE_NAME
from .babi_eval import babi_eval
from .cube import Cube

RETENTION_DAYS = config.getint('babi', 'retention_days', default=30)
VALID_FIRST_SYMBOLS = '_abcdefghijklmnopqrstuvwxyz'
VALID_NEXT_SYMBOLS = '_0123456789'
VALID_SYMBOLS = VALID_FIRST_SYMBOLS + VALID_NEXT_SYMBOLS

RIGHT_ARROW_HTML = '&#x27A1'
RIGHT_ARROW = html.unescape(RIGHT_ARROW_HTML)
DOWN_ARROW_HTML = '&#x2B07'
DOWN_ARROW = html.unescape(DOWN_ARROW_HTML)
MEASURE_HTML = '&#x1F4CA'
MEASURE = html.unescape(MEASURE_HTML)
PROPERTY_HTML = '&#x1F6C8'
PROPERTY = html.unescape(PROPERTY_HTML)
FIRE_HTML = '&#x1F525;'
FIRE = html.unescape(FIRE_HTML)
TICK_HTML = '&#x2705;'
TICK = html.unescape(TICK_HTML)
HOURGLASS_HTML = '&#x231B;'
HOURGLASS = html.unescape(HOURGLASS_HTML)

logger = logging.getLogger(__name__)

def save_virtual_workbook(workbook):
    with tempfile.NamedTemporaryFile() as tmp:
        save_workbook(workbook, tmp.name)
        with open(tmp.name, 'rb') as f:
            return f.read()

def convert_to_symbol(text):
    if not text:
        return 'x'
    text = text.replace('%', 'percent')
    text = unidecode.unidecode(text)
    text = text.lower()
    if text[0] not in VALID_FIRST_SYMBOLS:
        symbol = '_'
    else:
        symbol = ''
    for x in text:
        if not x in VALID_SYMBOLS:
            if symbol[-1] == '_':
                continue
            symbol += '_'
        else:
            symbol += x

    if len(symbol) > 1 and symbol[-1] == '_':
        symbol = symbol[:-1]
    return symbol

def generate_html_table(records):
    table = "<table>"
    tag = 'th'
    for row in records:
        table += "<tr>"
        for cell in row:
            align = 'right'
            if cell is None:
                cell = '<i>NULL</i>'
            elif isinstance(cell, datetime):
                cell = cell.strftime('%Y-%m-%d %H:%M:%S')
            elif isinstance(cell, mdatetime.date):
                cell = cell.strftime('%Y-%m-%d')
            elif isinstance(cell, mdatetime.time):
                cell = cell.strftime('%H:%M:%S')
            elif isinstance(cell, (float, int)):
                cell = str(cell)
            else:
                cell = str(cell)
                align = 'left'
            table += f"<{tag} align='{align}'>{cell}</{tag}>"
        tag = 'td'
        table += "</tr>"
    table += "</table>"
    return table


class CircularDependencyError(Exception):
    pass


class Cluster(ModelSQL, ModelView):
    'BABI Table Cluster'
    __name__ = 'babi.table.cluster'
    name = fields.Char('Name', required=True, readonly=True)
    active = fields.Function(fields.Boolean('Active'), 'get_active',
        searcher="search_active")
    tables = fields.One2Many('babi.table', 'cluster', 'Tables')
    root_tables = fields.Function(fields.One2Many('babi.table', None,
        'Tables'), 'get_root_tables')
    computation_start_date = fields.DateTime('Computation Start Date', readonly=True)
    computation_end_date = fields.DateTime('Computation End Date', readonly=True)
    elapsed = fields.Function(fields.TimeDelta('Elapsed'), 'get_elapsed')
    crons = fields.One2Many('ir.cron', 'babi_cluster', 'Schedulers', domain=[
            ('method', '=', 'babi.table.cluster|_compute')
            ], context={
            'babi_cluster': Eval('id', -1),
            }, depends=['id'])
    computation_state = fields.Selection([
            (None, 'Not Computed'),
            ('computing', HOURGLASS + ' Computing'),
            ('failed', FIRE + ' Failed'),
            ('successful', TICK + ' Successful'),
            ], 'State', readonly=True)

    @classmethod
    def __setup__(cls):
        super().__setup__()
        cls._order.insert(0, ('name', 'ASC'))
        cls._buttons.update({
                'compute': {},
                })

    @classmethod
    def copy(cls, clusters, default=None):
        if default is None:
            default = {}
        default = default.copy()
        default.setdefault('tables')
        return super().copy(clusters, default=default)

    def get_active(self, name):
        return bool(self.tables)

    @classmethod
    def search_active(cls, name, clause):
        if (clause[1] == '=' and clause[2]
                or clause[1] == '!=' and not clause[2]):
            domain = [('tables', '!=', None)]
        else:
            domain = [('tables', '=', None)]
        return domain

    def get_elapsed(self, name):
        if not self.computation_start_date:
            return
        end = self.computation_end_date or datetime.now()
        elapsed = end - self.computation_start_date
        # round down to seconds
        elapsed = elapsed - mdatetime.timedelta(microseconds=elapsed.microseconds)
        return elapsed

    def get_root_tables(self, name):
        return [x for x in self.tables if not x.requires]

    @classmethod
    @ModelView.button
    def compute(cls, clusters, compute_warnings=False):
        Table = Pool().get('babi.table')

        with Transaction().set_context(queue_name=QUEUE_NAME):
            start = datetime.now()
            for cluster in clusters:
                tables = [x for x in cluster.tables if x.active]
                if not tables:
                    continue
                cluster.computation_start_date = start
                cluster.computation_end_date = None
                cluster.computation_state = 'computing'
                cluster.save()
                Table.compute_cluster(tables)

            cls.save(clusters)

    @classmethod
    def get_compute_order(cls, tables):
        """
        Return the order in which tables should be computed to satisfy
        dependencies.

        :param tables: Dictionary where keys are tables and values are their
        dependencies.
        :return: List with the creation order of tables.
        """
        # Calculate the in-degree of each table
        in_degree = {table: 0 for table in tables}
        for dependents in tables.values():
            for dependent in dependents:
                in_degree[dependent] += 1

        # Add tables with no dependencies (in-degree 0) to the process
        # Sort them to ensure that the order is always the same for tables in the 0
        # degree
        queue = sorted([table for table in in_degree if in_degree[table] == 0])

        order = []

        while queue:
            table = queue.pop(0)  # Take a table from the queue
            order.append(table)

            # Reduce the in-degree of tables that depend on this table
            # Again, sort the tables to ensure the order is always the same for the
            # tables in the same degree
            for dependent in sorted(tables[table]):
                in_degree[dependent] -= 1
                if in_degree[dependent] == 0:
                    queue.append(dependent)

        # Check for cycles (if not all tables are processed)
        if len(order) != len(tables):
            raise CircularDependencyError(' > '.join(order))

        return order


class TableUser(ModelSQL):
    'BABI Table User'
    __name__ = 'babi.table-res.user'
    babi_table = fields.Many2One('babi.table', 'Table', required=True,
        ondelete='CASCADE')
    user = fields.Many2One('res.user', 'User', required=True,
        ondelete='CASCADE')


class TableGroup(ModelSQL):
    'BABI Table Group'
    __name__ = 'babi.table-res.group'
    babi_table = fields.Many2One('babi.table', 'Table', required=True,
        ondelete='CASCADE')
    user = fields.Many2One('res.group', 'Group', required=True,
        ondelete='CASCADE')


class TableParameters(Model):
    "Table Parameters"
    __name__ = 'babi.table.parameters'

    @classmethod
    def __setup__(cls):
        super().__setup__()
        cls.__rpc__.update({
                'get_keys': RPC(instantiate=0),
                'search_get_keys': RPC(),
                })

    @classmethod
    def get_keys(cls, records):
        pool = Pool()
        Parameter = pool.get('babi.filter.parameter')

        keys = []
        records = Parameter.search([])
        for record in records:
            key = {
                'id': record.id,
                'name': record.name,
                'string': record.name,
                'help': None,
                'type': record.ttype,
                'domain': None,
                'sequence': None,
                }
            if record.ttype in ('float', 'numeric'):
                key['digits'] = (16, record.digits)
            if record.ttype == 'many2one':
                key['type'] = 'integer'
            elif record.ttype == 'many2many':
                key['type'] = 'char'
            keys.append(key)
        return keys

    @classmethod
    def search_get_keys(cls, domain, limit=None):
        return cls.get_keys([])


class Table(DeactivableMixin, ModelSQL, ModelView):
    'BABI Table'
    __name__ = 'babi.table'
    name = fields.Char('Name', required=True)
    type = fields.Selection([
            (None, ''),
            ('model', 'Model'),
            ('table', 'Table'),
            ('view', 'View'),
            ], 'Type', required=True)
    internal_name = fields.Char('Internal Name', required=True)
    table_name = fields.Function(fields.Char('Table Name'),
        'on_change_with_table_name')
    model = fields.Many2One('ir.model', 'Model', states={
            'invisible': Eval('type') != 'model',
            'required': Eval('type') == 'model',
            }, domain=[('babi_enabled', '=', True)])
    filter = fields.Many2One('babi.filter', 'Filter', domain=[
            ('model', '=', Eval('model')),
            ], states={
            'invisible': Eval('type') != 'model',
            })
    fields_ = fields.One2Many('babi.field', 'table', 'Fields')
    query = fields.Text('Query', states={
            'invisible': ~Eval('type').in_(['table', 'view']),
            }, depends=['type'])
    timeout = fields.Integer('Timeout', required=True, help='If table '
        'calculation should take more than the specified timeout (in seconds) '
        'the process will be stopped automatically.')
    preview_limit = fields.Integer('Preview Limit', required=True)
    preview = fields.Function(fields.Binary('Preview',
        filename='preview_filename'), 'get_preview')
    preview_filename = fields.Function(fields.Char('Preview Filename'),
        'get_preview_filename')
    babi_raise_user_error = fields.Boolean('Raise User Error',
        help='Will raise a UserError in case of an error in the table.')
    compute_error = fields.Text('Compute Error', states={
            'invisible': ~Bool(Eval('compute_error')),
            }, readonly=True)
    compute_warning_error = fields.Text('Compute Warning Error', states={
            'invisible': ~Bool(Eval('compute_warning_error')),
            }, readonly=True,)
    crons = fields.One2Many('ir.cron', 'babi_table', 'Schedulers', domain=[
            ('method', '=', 'babi.table|_compute')
            ], context={
            'babi_table': Eval('id', -1),
            }, depends=['id'])
    cluster = fields.Many2One('babi.table.cluster', 'Cluster',
        ondelete='SET NULL')
    requires = fields.One2Many('babi.table.dependency', 'required_by',
        'Requires', readonly=True)
    required_by = fields.One2Many('babi.table.dependency', 'table',
        'Required By', readonly=True)
    requires_tables = fields.Function(fields.One2Many('babi.table', None,
        'Requires Tables'), 'get_requires_tables')
    required_by_tables = fields.Function(fields.One2Many('babi.table', None,
        'Required By Tables'), 'get_required_by_tables')
    ai_request = fields.Text('AI Request', states={
            'invisible': Eval('type') == 'model',
            })
    ai_response = fields.Text('AI Response', readonly=True, states={
            'invisible': Eval('type') == 'model',
            })
    warn = fields.Selection([
            (None, 'Never'),
            ('records', 'Records Found'),
            ('no-records', 'No Records Found'),
            ('always', 'Always'),
            ], 'Warn')
    email_template = fields.Many2One('electronic.mail.template',
        'Email Template', domain=[('model.name', '=', 'babi.warning')],
        states={
            'invisible': ~Bool(Eval('warn')),
            })
    warning_description = fields.Text('Description', states={
            'invisible': ~Bool(Eval('warn')),
            })
    cluster_date = fields.DateTime('Request Date', readonly=True)
    calculation_date = fields.DateTime('Date of calculation', readonly=True)
    calculation_time = fields.Float('Time taken to calculate (in seconds)',
        digits=(16, 6), readonly=True)
    last_warning_execution = fields.DateTime('Last Warning Execution',
        readonly=True)
    related_field = fields.Many2One('babi.field', 'Related Field', domain=[
            ('table.id', '=', Eval('id', -1)),
            ], states={
            'invisible': ~Bool(Eval('warn'))
            },ondelete='SET NULL')
    related_model = fields.Many2One('ir.model', 'Related Model', states={
            'required': Bool(Eval('related_field')),
            'invisible': ~Bool(Eval('warn'))
            })
    group = fields.Many2One('res.group', 'Group', ondelete='SET NULL', states={
            'invisible': ~Bool(Eval('warn'))
            })
    user = fields.Many2One('res.user', 'User', ondelete='SET NULL', states={
            'invisible': Bool(Eval('user_field')) | ~Bool(Eval('warn')),
            })
    user_field = fields.Many2One('babi.field', 'User Field', domain=[
            ('table.id', '=', Eval('id', -1)),
            ], ondelete='SET NULL', states={
            'invisible': Bool(Eval('user')) | ~Bool(Eval('warn')),
            })
    employee = fields.Many2One('company.employee', 'Employee',
        ondelete='SET NULL', states={
            'invisible': Bool(Eval('employee_field')) | ~Bool(Eval('warn')),
            })
    employee_field = fields.Many2One('babi.field', 'Employee Field', domain=[
            ('table.id', '=', Eval('id', -1)),
            ], ondelete='SET NULL', states={
            'invisible': Bool(Eval('employee')) | ~Bool(Eval('warn')),
            })
    company = fields.Many2One('company.company', 'Company', ondelete='SET NULL',
        states={
        'invisible': Bool(Eval('company_field')) | ~Bool(Eval('warn')),
        })
    company_field = fields.Many2One('babi.field', 'Company Field', domain=[
            ('table.id', '=', Eval('id', -1)),
            ], ondelete='SET NULL', states={
            'invisible': Bool(Eval('company')) | ~Bool(Eval('warn')),
            })
    party = fields.Many2One('party.party', 'Party', ondelete='SET NULL',
        states={
        'invisible': Bool(Eval('party_field')) | ~Bool(Eval('warn')),
        },
        context={
            'company': Eval('company', -1),
            },
        depends={'company'})
    party_field = fields.Many2One('babi.field', 'Party Field', domain=[
            ('table.id', '=', Eval('id', -1)),
            ], ondelete='SET NULL', states={
            'invisible': Bool(Eval('party')) | ~Bool(Eval('warn')),
            })
    pivot_table = fields.Function(fields.Char('Pivot Table'), 'get_pivot_table')
    access_users = fields.Many2Many('babi.table-res.user', 'babi_table', 'user',
        'Access Users')
    access_groups = fields.Many2Many('babi.table-res.group', 'babi_table', 'user',
        'Access Groups')
    pivots = fields.One2Many('babi.pivot', 'table', 'Pivot Tables')
    comment = fields.Text('Comment')
    parameters = fields.Dict('babi.table.parameters', 'Parameters', readonly=True,
        states={
            'invisible': ~Bool(Eval('parameters')),
            })

    @staticmethod
    def default_timeout():
        Config = Pool().get('babi.configuration')
        config = Config(1)
        return config.default_timeout or 30

    @staticmethod
    def default_preview_limit():
        return 10

    @classmethod
    def __setup__(cls):
        super().__setup__()
        cls._order.insert(0, ('name', 'ASC'))
        # Make internal_name unique
        t = cls.__table__()
        cls._sql_constraints = [
            ('internal_name_exclude', Exclude(t, (t.internal_name, Equal),
                    where=(t.active == Literal(True))
                    & (t.internal_name != '')),
                'babi.msg_table_internal_name_unique'),
            ]

        cls._buttons.update({
                'ai': {},
                'compute': {},
                'compute_warning': {
                    'invisible': ~Bool(Eval('warn')),
                    },
                'clear_cache': {},
                })

    @classmethod
    def __register__(cls, module_name):
        super().__register__(module_name)
        cursor = Transaction().connection.cursor()
        sql_table = cls.__table__()

        # Migration to 7.2: rename query to view
        cursor.execute(*sql_table.update([sql_table.type], ['view'],
                where=sql_table.type == 'query'))

    @classmethod
    def view_attributes(cls):
        return super().view_attributes() + [
            ('//page[@id="warning"]', 'states', {
                    'invisible': ~Bool(Eval('warn')),
                    })
            ]

    @classmethod
    def create(cls, vlist):
        tables = super().create(vlist)
        for table in tables:
            table.update_table_dependencies()
        return tables

    @classmethod
    def write(cls, *args):
        args = [x.copy() for x in args]
        actions = iter(args)
        for tables, values in zip(actions, actions):
            if 'ai_request' in values and not 'ai_response' in values:
                values['ai_response'] = None
            if 'internal_name' not in values:
                continue
            with Transaction().set_context(queue_name=QUEUE_NAME):
                for table in tables:
                    cls.__queue__._drop(table)

        super().write(*args)

        actions = iter(args)
        for tables, values in zip(actions, actions):
            for table in tables:
                if tuple(values.keys()) != ('cluster',):
                    table.update_table_dependencies()
                if 'internal_name' in values:
                    with Transaction().set_context(queue_name=QUEUE_NAME):
                        cls.__queue__._drop(table)

    @classmethod
    def delete(cls, tables):
        for table in tables:
            cls.__queue__._drop(table)
        super().delete(tables)

    @classmethod
    def clean(cls, date=None):
        if date is None:
            date = datetime.now() - timedelta(days=RETENTION_DAYS)
        tables = cls.search([
                ('create_date', '<', date),
                ('parameters', '!=', None),
                ])
        # We need to use the queue to delete the tables because if we try to
        # delete during a cron call, it will wait forever given that a ir.cron
        # record is being locked by the process and the ir.cron table has a m2o
        # to the babi.table table.
        cls.__queue__.delete(tables)

    @classmethod
    def copy(cls, tables, default=None):
        pool = Pool()
        Company = pool.get('company.company')
        Pivot = pool.get('babi.pivot')
        Lang = pool.get('ir.lang')

        if default is None:
            default = {}

        locale = Transaction().context.get(
            'report_lang', Transaction().language).split('_')[0]
        lang, = Lang.search([
                ('code', '=', locale or 'en'),
                ])

        now = datetime.now()
        company_id = Transaction().context.get('company')
        if company_id:
            company = Company(company_id)
            if company.timezone:
                timezone = pytz.timezone(company.timezone)
                now = timezone.localize(now)
                now = now + now.utcoffset()
        now = lang.strftime(now)


        default = default.copy()
        default.setdefault('name', lambda x: x['name'] + f' ({now})')
        default.setdefault('internal_name', lambda x: (
                convert_to_symbol(x['name'] + f' ({now})')))
        default.setdefault('pivots')
        default.setdefault('related_field')
        default.setdefault('user_field')
        default.setdefault('employee_field')
        default.setdefault('company_field')
        new_tables = super().copy(tables, default=default)
        to_save = []
        for old, new in zip(tables, new_tables):
            rel = {x.internal_name: x for x in new.fields_}
            new.related_field = (old.related_field
                and rel.get(old.related_field.internal_name))
            new.user_field = (old.user_field
                and rel.get(old.user_field.internal_name))
            new.employee_field = (old.employee_field
                and rel.get(old.employee_field.internal_name))
            new.company_field = (old.company_field
                and rel.get(old.company_field.internal_name))
            to_save.append(new)
            Pivot.copy(old.pivots, default={
                    'table': new,
                    })

        cls.save(to_save)
        return new_tables

    @fields.depends('user_field')
    def on_change_user(self):
        if self.user_field:
            self.user = None

    @fields.depends('user')
    def on_change_user_field(self):
        if self.user:
            self.user_field = None

    @fields.depends('employee_field')
    def on_change_employee(self):
        if self.employee_field:
            self.employee = None

    @fields.depends('employee')
    def on_change_employee_field(self):
        if self.employee:
            self.employee_field = None

    @fields.depends('company_field')
    def on_change_company(self):
        if self.company_field:
            self.company = None

    @fields.depends('company')
    def on_change_company_field(self):
        if self.company:
            self.company_field = None

    def get_rec_name(self, name):
        name = self.name
        if self.compute_error:
            name = FIRE + ' ' + name
        return name

    def update_table_dependencies(self):
        pool = Pool()
        Dependency = pool.get('babi.table.dependency')
        Cluster = pool.get('babi.table.cluster')

        Dependency.delete(self.requires)
        Dependency.delete(self.required_by)

        tables = {x.table_name: x for x in self.search([])}
        to_save = []

        required_tables = self.get_required_table_names()
        for name in required_tables:
            dependency = Dependency()
            dependency.required_by = self
            dependency.name = name
            dependency.table = tables.get(name)
            to_save.append(dependency)

        requiredby_tables = self.get_required_by_table_names() - required_tables
        for name in requiredby_tables:
            dependency = Dependency()
            dependency.required_by = tables.get(name)
            dependency.name = self.table_name
            dependency.table = self
            to_save.append(dependency)

        Dependency.save(to_save)

        cluster = self.get_cluster()
        order = self.get_compute_order(cluster)
        if len(order) <= 1:
            if self.cluster:
                self.cluster = None
                self.save()
                for table in cluster.tables:
                    table.update_table_dependencies()
            return
        table = order[0]
        # TODO: We should handle the case when we need to rename the cluster
        clusters = Cluster.search([('name', '=', table.table_name)], limit=1)
        if clusters:
            cluster = clusters[0]
        else:
            cluster = Cluster()
            cluster.name = table.table_name
            cluster.save()
        to_save = []
        for table in order:
            if table.cluster != cluster:
                table.cluster = cluster
                to_save.append(table)
        self.__class__.save(to_save)

    def get_requires_tables(self, name):
        return [x.table for x in self.requires if x.table]

    def get_required_by_tables(self, name):
        return [x.required_by for x in self.required_by if x.required_by]

    def get_required_table_names(self):
        if self.type and self.type == 'model':
            return set()
        query = self.query or ''
        tables = {x for x in query.split() if x.startswith('__')}
        return tables

    def get_required_by_table_names(self):
        tables = self.search([
                ('type', 'in', ['table', 'view']),
                ('query', 'ilike', '%' + self.table_name + '%'),
                ])
        res = set()
        for table in tables:
            if self.table_name in table.get_required_table_names():
                res.add(table.table_name)
        return res

    def get_records(self):
        table = []
        table.append([x.internal_name for x in self.fields_])
        try:
            table += self.execute_query()
        except Exception as e:
            raise UserError(gettext('babi.msg_error_obtaining_records',
                    table=self.rec_name, error=str(e)))
        return table

    def get_object_records(self):
        records = self.get_records()
        fields = records[0]
        records = records[1:]
        return [SimpleNamespace(**dict(zip(fields, x))) for x in records]

    def get_html(self, limit=None):
        start = time.time()
        content = None
        try:
            records = self.execute_query(limit=limit)
        except Exception as e:
            content = str(e)

        elapsed = time.time() - start
        if not content:
            table = []
            row = [x.internal_name for x in self.fields_]
            table.append(row)
            for record in records:
                table.append(record)
            content = '%(table)s<br/>%(elapsed).2fms' % {
                'table': generate_html_table(table),
                'elapsed': elapsed * 1000,
                }

        html = '''<!DOCTYPE html>
             <html>
             <head>
             <style>
             table, th, td {
                border: 1px solid black;
                border-collapse: collapse;
                padding: 5px;
             }
             * {
                font-family: monospace;
             }
             </style>
             </head>
             <body>%s</body></html>
        ''' % content
        return html

    def get_preview(self, name):
        return self.get_html(self.preview_limit).encode()

    def get_preview_filename(self, name):
        return self.internal_name + '.html'

    @classmethod
    def validate(cls, tables):
        super().validate(tables)
        for table in tables:
            table.check_internal_name()

    def check_internal_name(self):
        if not self.internal_name[0] in VALID_FIRST_SYMBOLS:
            raise UserError(gettext(
                'babi.msg_invalid_table_internal_name_first_character',
                table=self.rec_name, internal_name=self.internal_name))
        for symbol in self.internal_name:
            if not symbol in VALID_SYMBOLS:
                raise UserError(gettext('babi.msg_invalid_table_internal_name',
                        table=self.rec_name, internal_name=self.internal_name))

    @fields.depends('name')
    def on_change_name(self):
        self.internal_name = convert_to_symbol(self.name)

    def replace_parameters(self, expression):
        if not self.filter or not self.filter.parameters or not self.parameters:
            return expression
        try:
            if '%(' in expression:
                expression = expression % self.parameters
            else:
                expression = expression.format(**self.parameters)
        except KeyError as message:
            if self.report.babi_raise_user_error:
                raise UserError(
                    gettext('babi.invalid_parameters',
                    key=str(message)))
            raise
        return expression

    def get_python_filter(self):
        if self.filter and self.filter.python_expression:
            return self.replace_parameters(self.filter.python_expression)

    def get_domain_filter(self):
        domain = '[]'
        if self.filter and self.filter.domain:
            domain = self.filter.domain
            if '__' in domain:
                domain = str(PYSONDecoder().decode(domain))
        domain = self.replace_parameters(domain)
        return eval(domain, {
                'datetime': mdatetime,
                'false': False,
                'true': True,
                })

    def get_context(self):
        pool = Pool()
        Date = pool.get('ir.date')
        if not self.filter or not self.filter.context:
            return
        context = self.replace_parameters(self.filter.context)
        ev = EvalWithCompoundTypes(names={}, functions={
            'date': lambda x: datetime.strptime(x, '%Y-%m-%d').date(),
            'datetime': lambda x: datetime.strptime(x, '%Y-%m-%d'),
            'today': Date.today(),
            })
        context = ev.eval(context)
        return context

    def get_pivot_table(self, name):
        database = Transaction().database.name
        return http_host() + urllib.parse.quote(
            f'/{database}/babi/pivot/{self.table_name}/null')

    @property
    def ai_sql_tables(self):
        return {'account_invoice', 'account_invoice_line'}

    @classmethod
    @ModelView.button
    def csv(cls, tables):
        for table in tables:
            try:
                records = table.execute_query()
            except Exception as e:
                raise UserError(gettext('babi.msg_table_csv_error',
                        table=table.rec_name, error=str(e)))
            filename = table.internal_name + '.csv'
            with open(filename, 'w', newline='') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerows(records)


    @classmethod
    @ModelView.button
    def ai(cls, tables):
        cursor = Transaction().connection.cursor()

        from openai import OpenAI
        client = OpenAI(
            organization=config.get('openai', 'organization'),
            api_key=config.get('openai', 'api_key')
            )
        for table in tables:
            sqltables = dict.fromkeys(table.ai_sql_tables)

            t = sql.Table('columns', schema='information_schema')
            query = t.select(t.table_name, t.column_name,
                t.data_type)
            query.where = (t.table_schema == 'public') & \
                (t.table_name.in_(tuple(sqltables.keys())))
            cursor.execute(*query)
            for table_name, column_name, data_type in cursor.fetchall():
                sqltables[table_name] = sqltables[table_name] or {}
                sqltables[table_name][column_name] = data_type

            request = '''Given the following tables:

            %s

            Write an SQL query that returns the following information:

            %s

            Query:''' % (json.dumps(sqltables), table.ai_request)
            messages = [{
                'role': 'system',
                'content': 'Always return an SQL query that is suitable for PostgreSQL',
                }, {
                'role': 'user',
                'content': request,
                }]
            response = client.chat.completions.create(model="gpt-3.5-turbo",
            messages=messages)
            if response.choices:
                query = response.choices[0].message.content
                if not table.query:
                    table.query = query
                table.ai_response = query
                table.save()

    @classmethod
    @ModelView.button
    def compute(cls, tables):
        pool = Pool()
        ModelData = pool.get('ir.model.data')
        Action = pool.get('ir.action')

        for table in tables:
            if (table.filter and table.filter.parameters
                    and not table.parameters):
                if len(tables) > 1:
                    raise UserError(gettext('babi.msg_table_parametrize',
                            table=table.rec_name))
                # Clear table data that may exist if the table did not have
                # parameters before
                table.compute_error = None
                table.compute_warning_error = None
                table.calculation_date = None
                table.calculation_time = None
                table.save()
                table.clear_cache([table])
                table._drop()
                action_id = Action.get_action_id(ModelData.get_id('babi',
                        'table_parametrize_wizard'))
                values = Action(action_id).get_action_value()
                return values

        with Transaction().set_context(queue_name=QUEUE_NAME):
            for table in tables:
                if not table.active:
                    continue
                table.cluster_date = None
                cls.__queue__._compute(table)
            cls.save(tables)

    def get_cluster(self, tables=None):
        if tables is None:
            tables = {self}
        for dep in self.requires:
            table = dep.table
            if not table or table in tables:
                continue
            tables.add(table)
            table.get_cluster(tables)
        for dep in self.required_by:
            table = dep.required_by
            if table in tables:
                continue
            tables.add(table)
            table.get_cluster(tables)
        # Return the tables sorted by name so the order is always the same no
        # matter from which table we get the cluster
        return sorted(tables, key=lambda x: x.name)

    def get_compute_order(self, cluster=None):
        Cluster = Pool().get('babi.table.cluster')

        tables = {}
        mapping = {}
        if cluster is None:
            cluster = self.get_cluster()
        for table in cluster:
            mapping[table.table_name] = table
            tables[table.table_name] = [x.required_by.table_name for x in table.required_by]
        order = Cluster.get_compute_order(tables)
        return [mapping[x] for x in order]

    def get_next_in_cluster(self):
        order = self.get_compute_order()
        if not self in order:
            return
        idx = order.index(self) + 1
        if idx >= len(order):
            return
        return order[idx]

    @classmethod
    def compute_cluster(cls, tables, compute_warnings=False):
        with Transaction().set_context(queue_name=QUEUE_NAME):
            all_ = set()
            to_save = []
            for table in tables:
                if table in all_:
                    continue
                cluster = table.get_cluster()
                try:
                    order = table.get_compute_order()
                except CircularDependencyError as e:
                    raise UserError(gettext('babi.msg_circular_dependency',
                            sequence=e.message))
                first = order[0]
                first.cluster_date = datetime.now()
                to_save.append(first)
                cls.__queue__._compute(first,
                    compute_warnings=compute_warnings, cluster=True)
                all_ |= set(cluster)
            cls.save(to_save)

    @classmethod
    @ModelView.button
    def compute_warning(cls, tables):
        with Transaction().set_context(queue_name=QUEUE_NAME):
            for table in tables:
                if not table.active:
                    continue
                cls.__queue__.compute_warnings(table)

    @classmethod
    @ModelView.button
    def clear_cache(cls, tables):
        for table in tables:
            cube = Cube(table.table_name)
            cube.clear_cache()

        # Remove all cache tables that are not in use by any of the tables
        tables = cls.search([])
        Cube.clear_orphan_caches([x.table_name for x in tables])

    @fields.depends('internal_name')
    def on_change_with_table_name(self, name=None):
        if not self.internal_name:
            return
        # Add a suffix to the table name to prevent removing production tables
        return '__' + self.internal_name

    def get_query(self, fields=None, where=None, groupby=None, limit=None):
        query = 'SELECT '
        if fields:
            query += ', '.join(fields) + ' '
        else:
            query += '* '

        if self.type == 'view':
            query += 'FROM (%s) AS a ' % self._stripped_query
        else:
            query += 'FROM %s ' % self.table_name
        if where:
            where = where.format(**Transaction().context)
            query += 'WHERE %s ' % where

        if groupby:
            query += 'GROUP BY %s ' % ', '.join(groupby) + ' '

        if fields:
            query += 'ORDER BY %s' % ', '.join(fields)

        if limit:
            query += ' LIMIT %d' % limit
        return query

    def execute_query(self, fields=None, where=None, groupby=None, timeout=None,
            limit=None):
        if (self.type != 'view'
                and not backend.TableHandler.table_exist(self.table_name)):
            return []
        with Transaction().new_transaction() as transaction:
            cursor = transaction.connection.cursor()
            self._set_statement_timeout(timeout)
            query = self.get_query(fields, where=where, groupby=groupby,
                limit=limit)
            cursor.execute(query)
            records = cursor.fetchall()
            self._reset_statement_timeout()
        if self.type == 'model':
            # Sort records based on the order and descending fields from fields_
            headers = fields
            if not headers:
                headers = [x.internal_name for x in self.fields_]
            sort_key = []
            for field in self.fields_:
                if not field.order:
                    continue
                if not field.internal_name in headers:
                    continue
                sort_key.append((field.order,
                        headers.index(field.internal_name),
                        field.descending or False))
            sort_key.sort()
            if sort_key:
                # Sort on reverse order taking advantage of the stability of
                # the sort method
                for item in reversed(sort_key):
                    records.sort(key=lambda x: x[item[1]], reverse=item[2])
        return records

    def timeout_exception(self):
        raise TimeoutException

    def _compute(self, compute_warnings=False, cluster=False):
        self.clear_cache([self])
        start_time = time.time()
        try:
            if self.type == 'model':
                if not self.fields_:
                    raise UserError(gettext('babi.msg_table_no_fields',
                            table=self.name))
                self._compute_model()
            elif self.type == 'table':
                self._compute_table()
            elif self.type == 'view':
                self._compute_view()

        except Exception as e:
            # In case there is a create view error or SQL typo,
            # we do rollback to obtain a value from the gettext()
            Transaction().connection.rollback()
            notify(gettext('babi.msg_table_failed', table=self.rec_name))
            self.compute_error = f'{e}\n{traceback.format_exc()}'
            self.save()
            if cluster and self.cluster:
                self.cluster.computation_end_date = datetime.now()
                self.cluster.computation_state = 'failed'
                self.cluster.save()
            return

        if cluster and self.cluster:
            next_ = self.get_next_in_cluster()
            if next_:
                next_.cluster_date = self.cluster_date
                next_.save()
                self.__class__.__queue__._compute(next_,
                    compute_warnings=compute_warnings, cluster=cluster)
            else:
                self.cluster.computation_end_date = datetime.now()
                self.cluster.computation_state = 'successful'
                self.cluster.save()

        self.compute_error = None
        self.compute_warning_error = None
        end_time = time.time()
        self.save()
        notify(gettext('babi.msg_table_successful', table=self.rec_name))
        self.calculation_date = datetime.now()
        self.calculation_time = round(end_time - start_time,
            self.__class__.calculation_time.digits[1])
        self.save()
        if compute_warnings:
            self.__class__.__queue__.compute_warnings(self)

    def compute_warnings(self):
        self.compute_warning_error = None
        self.save()
        query = self.get_query()
        if query:
            user_id = None
            if self.user_field:
                user_id = self.user_field.internal_name
            elif self.user:
                user_id = self.user.id
            else:
                user_id = 'NULL'

            employee_id = None
            if self.employee_field:
                employee_id = self.employee_field.internal_name
            elif self.employee:
                employee_id = self.employee.id
            else:
                employee_id = 'NULL'

            company_id = None
            if self.company_field:
                company_id = self.company_field.internal_name
            elif self.company:
                company_id = self.company.id
            else:
                company_id = 'NULL'

            if self.party_field:
                party_id = self.party_field.internal_name
            elif self.party:
                party_id = self.party.id
            else:
                party_id = 'NULL'

            query_full = 'SELECT '
            query_full += '   count(*), '
            query_full += f'  {user_id} AS user_id, '
            query_full += f'  {employee_id} as employee_id, '
            query_full += f'  {party_id} as party_id, '
            query_full += f'  {company_id} as company_id '
            query_full += 'FROM (%s) AS compute_warnings_subquery ' % query

            group_by = [user_id, employee_id, company_id, party_id]
            group_by = [x for x in group_by
                if isinstance(x, str) and x != 'NULL']
            if group_by:
                query_full += 'GROUP BY ' + ', '.join(group_by)

        to_create = []
        self.last_warning_execution = datetime.now()
        cursor = Transaction().connection.cursor()
        cursor.execute(query_full)
        query_last = cursor.fetchall()
        for x in query_last:
            count = x[0]
            if (self.warn == 'always'
                    or (self.warn == 'records' and count)
                    or (self.warn == 'no-records' and not count)):
                to_create.append({
                        'timestamp': self.last_warning_execution,
                        'table': self.id,
                        'count': count,
                        'user': x[1],
                        'employee': x[2],
                        'party': x[3],
                        'company': x[4],
                        'group': self.group,
                        })
        if to_create:
            try:
                warnings = Warning.create(to_create)
                for warning in warnings:
                    warning.send()
            except Exception as e:
                Transaction().connection.rollback()
                self.compute_warning_error = f'{e}\n{traceback.format_exc()}'
                self.save()

    def update_fields(self, field_names):
        pool = Pool()
        Field = pool.get('babi.field')

        # Update self.fields_
        to_save = []
        to_delete = []
        existing_fields = set([])
        for field in self.fields_:
            if field.internal_name not in field_names:
                to_delete.append(field)
                continue
            field.sequence = field_names.index(field.internal_name)
            existing_fields.add(field.internal_name)
            to_save.append(field)

        for field_name in (set(field_names) - existing_fields):
            field = Field()
            field.table = self
            field.name = field_name
            field.internal_name = field_name
            field.sequence = field_names.index(field.internal_name)
            to_save.append(field)

        Field.save(to_save)
        Field.delete(to_delete)

    @property
    def _stripped_query(self):
        if self.query:
            return self.query.strip().rstrip(';')
        else:
            return ''

    def _drop(self):
        # Given that the type may be changed from view to table and viceversa
        # we cannot rely on self.type to know if we have to execute DROP TABLE
        # or DROP VIEW.
        cursor = Transaction().connection.cursor()
        if backend.name != 'postgresql':
            cursor.execute('DROP TABLE IF EXISTS %s' % self.table_name)
            cursor.execute('DROP VIEW IF EXISTS %s' % self.table_name)
            return
        # In Postgres, trying to execute DROP VIEW on a TABLE will make
        # postgres complaint (even with the 'IF EXISTS' clause). And the same
        # will happen with DROP TABLE on a VIEW. So we must check if it exists
        # and its type.

        tables = sql.Table('tables', 'information_schema')
        cursor.execute(*tables.select(tables.table_type,
                where=(tables.table_name == self.table_name) &
                (tables.table_schema == 'public')))
        record = cursor.fetchone()
        if not record:
            return
        if record[0] == 'VIEW':
            cursor.execute(f'DROP VIEW IF EXISTS "{self.table_name}" CASCADE')
        else:
            cursor.execute(f'DROP TABLE IF EXISTS "{self.table_name}" CASCADE')

    def _set_statement_timeout(self, timeout=None):
        if backend.name != 'postgresql':
            return
        cursor = Transaction().connection.cursor()
        cursor.execute(f'SET statement_timeout TO {self.timeout * 1000};')

    def _reset_statement_timeout(self):
        if backend.name != 'postgresql':
            return
        cursor = Transaction().connection.cursor()
        cursor.execute('SET statement_timeout TO 0;')

    def _compute_view(self):
        with Transaction().new_transaction() as transaction:
            cursor = transaction.connection.cursor()
            # We must use a subquery because the _stripped_query may contain a
            # LIMIT clause
            self._set_statement_timeout()
            cursor.execute('SELECT * FROM (%s) AS subquery LIMIT 1' %
                self._stripped_query)
            self._reset_statement_timeout()

        field_names = [x[0] for x in cursor.description]
        self.update_fields(field_names)

        cursor = Transaction().connection.cursor()
        self._drop()
        cursor.execute('CREATE VIEW "%s" AS %s' % (self.table_name, self._stripped_query))

    def _compute_table(self):
        with Transaction().new_transaction() as transaction:
            self._drop()
            cursor = transaction.connection.cursor()
            self._set_statement_timeout()
            cursor.execute('CREATE TABLE "%s" AS %s' % (self.table_name,
                    self._stripped_query))
            self._reset_statement_timeout()
            cursor.execute('SELECT * FROM "%s" LIMIT 1' % self.table_name)

        field_names = [x[0] for x in cursor.description]
        self.update_fields(field_names)

    def _compute_model(self):
        Model = Pool().get(self.model.name)

        with Transaction().new_transaction() as transaction:
            cursor = transaction.connection.cursor()
            self._drop()
            fields = []
            for field in self.fields_:
                fields.append('"%s" %s' % (field.internal_name, field.sql_type()))
            cursor.execute('CREATE TABLE IF NOT EXISTS "%s" (%s);' % (
                    self.table_name, ', '.join(fields)))

            checker = TimeoutChecker(self.timeout, self.timeout_exception)
            domain = self.get_domain_filter()

            context = self.get_context()
            if not context:
                context = {}
            else:
                assert isinstance(context, dict)
            context['_datetime'] = None
            # This is needed when execute the wizard to calculate the report, to
            # ensure the company rule is used.
            context['_check_access'] = True

            python_filter = self.get_python_filter()

            table = sql.Table(self.table_name)
            columns = [sql.Column(table, x.internal_name) for x in self.fields_]
            expressions = [(x.expression.expression, x.expression.decimal_digits)
                for x in self.fields_]
            index = 0
            count = 0
            offset = 10000
            context['_record_cache_size'] = 10000

            with Transaction().set_context(**context):
                try:
                    records = Model.search(domain, offset=index * offset,
                        limit=offset)
                except Exception as message:
                    if self.babi_raise_user_error:
                        raise UserError(gettext(
                            'babi.create_data_exception',
                            error=repr(message)))
                    raise

            while records:
                checker.check()
                logger.info('Calculated %s, %s records in %s seconds'
                    % (self.model.name, count, checker.elapsed))

                to_insert = []
                for record in records:
                    if python_filter:
                        if not babi_eval(python_filter, record, convert_none=None):
                            continue
                    values = []
                    for expression, digits in expressions:
                        try:
                            values.append(babi_eval(expression, record,
                                    convert_none=None, digits=digits))
                        except Exception as message:
                            notify(gettext('babi.msg_compute_table_exception',
                                    table=self.name, field=field.name,
                                    record=record.id, error=repr(message)),
                                priority=1)
                            if self.babi_raise_user_error:
                                raise UserError(gettext(
                                    'babi.msg_compute_table_exception',
                                    table=self.name,
                                    field=field.name,
                                    record=record.id,
                                    error=repr(message)))
                            raise

                    to_insert.append(values)

                cursor.execute(*table.insert(columns=columns, values=to_insert))

                index += 1
                count += len(records)
                with Transaction().set_context(**context):
                    records = Model.search(domain, offset=index * offset,
                        limit=offset)

        logger.info('Calculated %s, %s records in %s seconds'
            % (self.model.name, count, checker.elapsed))

    def check_access(self, user=None):
        pool = Pool()
        User = pool.get('res.user')
        ModelData = pool.get('ir.model.data')


        if not user:
            user = User(Transaction().user)
        if not user:
            return False

        group_babi_admin = ModelData.get_id('babi', 'group_babi_admin')
        for group in user.groups:
            if group.id == group_babi_admin:
                return True
        if self.access_groups:
            for group in user.groups:
                if group in self.access_groups:
                    return True
        if self.access_users:
            if user not in self.access_users:
                return False
        return True


class Field(sequence_ordered(), ModelSQL, ModelView):
    'BABI Field'
    __name__ = 'babi.field'
    table = fields.Many2One('babi.table', 'Table', required=True,
        ondelete='CASCADE')
    name = fields.Char('Name', required=True)
    internal_name = fields.Char('Internal Name', required=True)
    expression = fields.Many2One('babi.expression', 'Expression', states={
            'invisible': Eval('table_type') != 'model',
            'required': Eval('table_type') == 'model'
            }, domain=[
            ('model', '=', Eval('model')),
            ])
    order = fields.Integer('Order', domain=[
        'OR',
        ('order', '>', 0),
        ('order', '=', None),
        ], help='Positive integer. Indicates if the field must be sorted '
            'and in which position. Field with order 1 will be sorted before '
            '2, etc. Order is computed in memory so it can be slow with many '
            'records.')
    descending = fields.Boolean('Descending')
    model = fields.Function(fields.Many2One('ir.model', 'Model'),
        'on_change_with_model')
    type = fields.Function(fields.Selection(FIELD_TYPES, 'Type'),
        'on_change_with_type')
    table_type = fields.Function(fields.Selection([
            ('model', 'Model'),
            ('table', 'Table'),
            ('view', 'View'),
            ], 'Table Type'), 'on_change_with_table_type')

    @fields.depends('expression')
    def on_change_with_type(self, name=None):
        if self.expression:
            return self.expression.ttype

    @fields.depends('table', '_parent_table.type')
    def on_change_with_table_type(self, name=None):
        if self.table:
            return self.table.type

    @classmethod
    def __setup__(cls):
        super().__setup__()
        t = cls.__table__()
        cls._sql_constraints += [
            ('table_internal_name_uniq', Unique(t, t.table, t.internal_name),
                'Field must be unique per Table'),
            ]
        cls.__access__.add('table')

    @classmethod
    def validate(cls, babi_fields):
        super().validate(babi_fields)
        for babi_field in babi_fields:
            babi_field.check_internal_name()

    def sql_type(self):
        mapping = {
            'char': 'VARCHAR',
            'integer': 'INTEGER',
            'float': 'FLOAT',
            'numeric': 'NUMERIC',
            'boolean': 'BOOLEAN',
            'many2one': 'INTEGER',
            'date': 'DATE',
            'datetime': 'DATETIME',
            }
        return mapping[self.expression.ttype]

    def check_internal_name(self):
        if not self.internal_name[0] in VALID_FIRST_SYMBOLS:
            raise UserError(gettext('babi.msg_invalid_field_internal_name',
                    field=self.name, internal_name=self.internal_name))
        for symbol in self.internal_name:
            if not symbol in VALID_SYMBOLS:
                raise UserError(gettext('babi.msg_invalid_field_internal_name',
                        field=self.name, internal_name=self.internal_name))

    @fields.depends('name')
    def on_change_name(self):
        self.internal_name = convert_to_symbol(self.name)

    @fields.depends('name', 'expression', methods=['on_change_name'])
    def on_change_expression(self):
        if self.expression:
            self.name = self.expression.name
            self.on_change_name()

    @fields.depends('table', '_parent_table.model')
    def on_change_with_model(self, name=None):
        if self.table and self.table.model:
            return self.table.model.id


class TableDependency(ModelSQL, ModelView):
    'BABI Table Dependency'
    __name__ = 'babi.table.dependency'
    required_by = fields.Many2One('babi.table', 'Required By', required=True,
        ondelete='CASCADE')
    name = fields.Char('Name')
    table = fields.Many2One('babi.table', 'Requires', ondelete='SET NULL')

    @classmethod
    def __setup__(cls):
        super().__setup__()
        cls.__access__.add('table')


class Warning(Workflow, ModelSQL, ModelView):
    'BABI Warning'
    __name__ = 'babi.warning'

    timestamp = fields.DateTime('Timestamp', required=True, readonly=True)
    table = fields.Many2One('babi.table', 'Table', required=True, readonly=True,
        ondelete='CASCADE')
    # We can have a warning to a user without acces to the table, we need to
    # use a function field to only show the table if the user have access to it.
    # A second "table" field must be used because there are certain fields that
    # depend on the original table, for example, the button that allows you to
    # see the records with warnings (to see the records with warnings you do
    # not need to have permissions on the table).
    table_to_show = fields.Function(
        fields.Many2One('babi.table', 'Table'), 'get_table_to_show')
    has_related_records = fields.Function(fields.Boolean('Has Related Field'),
            'get_has_related_records')
    count = fields.Integer('Records found', readonly=True, required=True)
    state = fields.Selection([
            ('pending', 'Pending'),
            ('done', 'Done'),
            ('ignored', 'Ignored'),
            ], 'State', readonly=True)
    company = fields.Many2One('company.company', 'Company', readonly=True,
            ondelete='CASCADE')
    employee = fields.Many2One('company.employee', 'Employee', readonly=True)
    description = fields.Function(fields.Text('Description', readonly=True),
            'get_description',)
    user = fields.Many2One('res.user', 'User', readonly=True,
            ondelete='CASCADE')
    group = fields.Many2One('res.group', 'Group', ondelete='CASCADE',
            readonly=True)
    party = fields.Many2One('party.party', 'Party', ondelete='CASCADE',
            readonly=True, context={
                'company': Eval('company', -1),
            }, depends=['company'])

    users = fields.Function(fields.Many2Many('res.user', None, None, 'Users'),
        'get_users')
    emails = fields.Function(fields.Char('E-mails'), 'get_emails')
    done_by = employee_field("Done By", states=['pending', 'done', 'ignored'])
    ignored_by = employee_field("Ignored By",
        states=['pending', 'done', 'ignored'])

    def get_rec_name(self, name):
        return f'{self.count} - {self.table.name}'

    @classmethod
    def search_rec_name(cls, name, clause):
        return [('table.name',) + tuple(clause[1:])]

    @classmethod
    def __setup__(cls):
        super().__setup__()
        cls._transitions |= set((
                ('pending', 'done'),
                ('pending', 'ignored'),
                ('done', 'ignored'),
                ('ignored', 'pending'),
                ))
        cls._buttons.update({
                'pending': {
                    'invisible': ((Eval('state') == 'done')
                        | (Eval('state') == 'pending')),
                    'depends': ['state'],
                    'icon': 'tryton-back'
                    },
                'do': {
                    'invisible': Eval('state') != 'pending',
                    'depends': ['state'],
                    'icon': 'tryton-ok',
                    },
                'ignore': {
                    'invisible': Eval('state') == 'ignored',
                    'depends': ['state'],
                    'icon': 'tryton-cancel',
                    },
                'open': {
                    'invisible': ~Bool(Eval('has_related_records')),
                    'depends': ['has_related_records'],
                    'icon': 'tryton-add',
                    },
                })

    @staticmethod
    def default_state():
        return 'pending'

    def get_has_related_records(self, name):
        if not self.table.related_field:
            return False
        return bool(self.count)

    def get_users(self, name):
        pool = Pool()
        User = pool.get('res.user')

        if self.user:
            return [self.user]
        if self.employee:
            return User.search([
                    ('employee.id', '=', self.employee.id),
                    ])
        if self.group:
            return [x for x in self.group.users if x.active]
        if self.company:
            return User.search([
                    ('companies.id', '=', self.company.id),
                    ])
        return User.search([])

    def get_emails(self, name):
        User = Pool().get('res.user')

        emails = []
        if self.user:
            emails = [self.user.email]
        elif self.employee:
            emails = [self.employee.party.email]
        elif self.group:
            users = User.search([('groups.id', '=', self.group.id)])
            emails = [user.email for user in users]
        elif self.company:
            users = User.search([('companies.id', '=', self.company.id)])
            emails = [user.email for user in users]
        elif self.party:
            emails = [self.party.email]
        else:
            users = User.search([])
            emails = [user.email for user in users]
        return ', '.join(sorted(list({x for x in emails if x})))

    @classmethod
    @ModelView.button
    @Workflow.transition('pending')
    @reset_employee('done_by', 'ignored_by')
    def pending(cls, warnings):
        pass

    @classmethod
    @ModelView.button
    @Workflow.transition('done')
    @set_employee('done_by')
    def do(cls, warnings):
        pass

    @classmethod
    @ModelView.button
    @Workflow.transition('ignored')
    @set_employee('ignored_by')
    def ignore(cls, warnings):
        pass

    def get_description(self, name):
        return self.table.warning_description

    @classmethod
    def get_table_to_show(cls, warnings, name):
        pool = Pool()
        Table = pool.get('babi.table')

        result = {}
        with Transaction().set_context(_check_access=True):
            tables = Table.search([
                ('id', 'in', [x.table.id for x in warnings])])
            tables_found = [t.id for t in tables]

            for warning in warnings:
                if warning.table.id in tables_found:
                    result[warning.id] = warning.table.id
                else:
                    result[warning.id] = None

        return result

    def get_ids(self):
        pool = Pool()
        Table = pool.get('babi.table')

        if not self.table.related_model:
            return []

        user_index = 1
        fields = [self.table.related_field.internal_name]
        if self.table.company_field:
            user_index += 1
            fields.append(self.table.company_field.internal_name)
        if self.table.user_field:
            fields.append(self.table.user_field.internal_name)
        if self.table.employee_field:
            fields.append(self.table.employee_field.internal_name)
        if self.table.party_field:
            fields.append(self.table.party_field.internal_name)
        records = Table.execute_query(self.table,
            fields=fields)

        if self.table.company_field and self.company:
            records = [x for x in records if x[1] == self.company.id]
        if self.table.user_field and self.user:
            records = [x for x in records if x[user_index] == self.user.id]
        if self.table.employee_field and self.employee:
            records = [x for x in records if x[user_index] == self.employee.id]
        if self.table.party_field and self.party:
            records = [x for x in records if x[user_index] == self.party.id]

        try:
            ids = []
            for record in records:
                value = record[0]
                if value is None:
                    continue
                # Allow using arrays (using array_agg) of ids or
                # just ids
                if isinstance(value, (tuple, list)):
                    ids += list([int(x) for x in value])
                else:
                    ids.append(int(value))
        except:
            raise UserError(gettext('babi.msg_not_converted',
                field=self.table.related_field.rec_name))
        return ids

    def get_records(self, ids=None):
        if not self.table.related_model:
            return []
        pool = Pool()
        Model = pool.get(self.table.related_model.name)
        if ids is None:
            ids = self.get_ids()
        return Model.search([('id', 'in', ids)])

    @classmethod
    @ModelView.button
    def open(cls, warnings):
        if len(set([x.table for x in warnings])) > 1:
            raise UserError(gettext('babi.msg_open_warning_single'))

        ids = []
        records = []
        for warning in warnings:
            wids = warning.get_ids()
            records += warning.get_records(wids)
            ids += wids

        if len(records) != len(ids):
            raise UserError(gettext('babi.msg_wrong_info',
                field=warning.table.related_field.rec_name,
                query_count=str(len(ids)),
                model_query_count=str(len(records))))

        return {
            'res_model': warning.table.related_model.name,
            'type': 'ir.action.act_window',
            'name': warning.table.related_model.name,
            'pyson_domain': f'[["id", "in", {ids}]]',
            'pyson_context': '{}',
            'pyson_order': '[]',
            'domains': [],
            }

    def send(self):
        if self.table.warn and self.table.email_template:
            self.table.email_template.render_and_send(
                self.table.email_template.id, [self])


class TableExcel(Report):
    'Table Excel Export'
    __name__ = 'babi.table.excel'

    @classmethod
    def execute(cls, ids, data):
        pool = Pool()
        Table = pool.get('babi.table')

        if not ids:
            return

        action, model = cls.get_action(data)
        cls.check_access(action, model, ids)

        def _convert_to_string(value):
            if isinstance(value, (Decimal, str, int, float, date, datetime,
                    dt_time, bool)):
                return value
            return str(value) if value is not None else None

        tables = Table.browse(ids)
        wb = Workbook()
        wb.remove(wb.active)
        for table in tables:
            ws = wb.create_sheet(table.name)
            for record in table.get_records():
                ws.append([_convert_to_string(item) for item in record])

        if len(tables) == 1:
            name = table.name
        else:
            name = gettext('babi.msg_tables_filename')
        return ('xlsx', save_virtual_workbook(wb), False, name)


class Pivot(ModelSQL, ModelView):
    'Pivot Table'
    __name__ = 'babi.pivot'

    table = fields.Many2One('babi.table', 'Table', required=True,
        ondelete='CASCADE')
    name = fields.Char('Name')
    active = fields.Function(fields.Boolean('Active'), 'get_active',
        searcher='search_active', setter='set_active')
    active_stored = fields.Boolean('Active Stored')
    row_dimensions = fields.One2Many('babi.pivot.row_dimension', 'pivot',
        'Row Dimensions')
    column_dimensions = fields.One2Many('babi.pivot.column_dimension', 'pivot',
        'Column Dimensions')
    measures = fields.One2Many('babi.pivot.measure', 'pivot', 'Measures')
    properties = fields.One2Many('babi.pivot.property', 'pivot', 'Properties')
    order = fields.One2Many('babi.pivot.order', 'pivot', 'Order')
    url = fields.Function(fields.Char('URL'), 'on_change_with_url')

    @staticmethod
    def default_active():
        return True

    @staticmethod
    def default_active_stored():
        return True

    def get_rec_name(self, name):
        res = []
        if self.name:
            res.append(self.name)
        if self.row_dimensions:
            item = RIGHT_ARROW + ' '
            item += ', '.join([x.field.rec_name for x in self.row_dimensions])
            res.append(item)
        if self.properties:
            item = PROPERTY + ' '
            item += ', '.join([x.field.rec_name for x in self.properties])
            res.append(item)
        if self.column_dimensions:
            item = DOWN_ARROW + ' '
            item += ', '.join([x.field.rec_name for x in self.column_dimensions])
            res.append(item)
        if self.measures:
            item = MEASURE + ' '
            item += ', '.join([x.field.rec_name for x in self.measures])
            res.append(item)
        return ' '.join(res)

    def get_active(self, name):
        return self.active_stored and self.table.active

    @classmethod
    def search_active(cls, name, clause):
        if ((clause[1] == '=' and clause[2])
                or (clause[1] == '!=' and clause[2])):
            domain = [
                ('active_stored', '=', True),
                ('table.active', '=', True),
                ]
        else:
            domain = [
                'OR',
                ('active_stored', '=', False),
                ('table.active', '=', False),
                ]
        return domain

    @classmethod
    def set_active(cls, pivots, name, value):
        for pivot in pivots:
            pivot.active_stored = value
        cls.save(pivots)

    @fields.depends('table', '_parent_table.pivot_table', methods=['get_cube'])
    def on_change_with_url(self, name=None):
        cube = self.get_cube()
        if not cube:
            return
        properties = cube.encode_properties()
        return f'{self.table.pivot_table.replace("/null", "")}/{properties}'

    @classmethod
    def copy(cls, pivots, default=None):
        pool = Pool()
        RowDimension = pool.get('babi.pivot.row_dimension')
        ColumnDimension = pool.get('babi.pivot.column_dimension')
        Measure = pool.get('babi.pivot.measure')
        Property = pool.get('babi.pivot.property')

        if default is None:
            default = {}
        else:
            default = default.copy()
        default.setdefault('row_dimensions', [])
        default.setdefault('column_dimensions', [])
        default.setdefault('measures', [])
        default.setdefault('properties', [])
        default.setdefault('order', [])
        new_pivots = super().copy(pivots, default)

        rd_to_save = []
        cd_to_save = []
        m_to_save = []
        p_to_save = []
        for old, new in zip(pivots, new_pivots):
            rel = {x.internal_name: x for x in new.table.fields_}
            for dimension in old.row_dimensions:
                record = RowDimension()
                record.pivot = new
                record.field = rel[dimension.field.internal_name]
                rd_to_save.append(record)
            for dimension in old.column_dimensions:
                record = ColumnDimension()
                record.pivot = new
                record.field = rel[dimension.field.internal_name]
                cd_to_save.append(record)
            for measure in old.measures:
                record = Measure()
                record.pivot = new
                record.field = rel[measure.field.internal_name]
                record.aggregate = measure.aggregate
                m_to_save.append(record)
            for property_ in old.properties:
                record = Property()
                record.pivot = new
                record.field = rel[property_.field.internal_name]
                p_to_save.append(record)
        RowDimension.save(rd_to_save)
        ColumnDimension.save(cd_to_save)
        Measure.save(m_to_save)
        Property.save(p_to_save)
        cls.update_order(new_pivots)
        return new_pivots

    @fields.depends('table', 'row_dimensions', 'column_dimensions', 'measures',
        'properties', 'order', '_parent_table.table_name')
    def get_cube(self):
        if not self.table:
            return
        order = []
        for item in self.order:
            if not item.element:
                continue
            if item.element.__name__ == 'babi.pivot.measure':
                order.append(((item.element.field.internal_name,
                            item.element.aggregate), item.order))
            else:
                order.append((item.element.field.internal_name, item.order))
        return Cube(table=self.table.table_name,
            rows=[x.field.internal_name for x in self.row_dimensions if x.field],
            columns=[x.field.internal_name for x in self.column_dimensions if x.field],
            measures=[(x.field.internal_name, x.aggregate)
                for x in self.measures if x.field and x.aggregate],
            properties=[x.field.internal_name for x in self.properties if x.field],
            order=order,
            )

    @classmethod
    def search_rec_name(cls, name, clause):
        return ['OR',
            ('name',) + tuple(clause[1:]),
            ('table.rec_name',) + tuple(clause[1:]),
            ('row_dimensions.field.rec_name',) + tuple(clause[1:]),
            ('column_dimensions.field.rec_name',) + tuple(clause[1:]),
            ('measures.field.rec_name',) + tuple(clause[1:]),
            ]

    @classmethod
    def __setup__(cls):
        super().__setup__()
        cls._order.insert(0, ('name', 'ASC'))
        cls.__access__.add('table')

    @classmethod
    def create(cls, vlist):
        pivots = super().create(vlist)
        cls.update_order(pivots)
        return pivots

    @classmethod
    def write(cls, *args):
        super().write(*args)
        cls.update_order(cls.browse([x[0] for x in args[::2]]))

    @classmethod
    def update_order(cls, pivots):
        pool = Pool()
        Order = pool.get('babi.pivot.order')

        to_save = []
        to_delete = set()
        for pivot in pivots:
            orders = [x.sequence for x in pivot.order if x.sequence is not None]
            if orders:
                sequence = max(orders) + 1
            else:
                sequence = 1

            records = []
            records += pivot.row_dimensions
            records += pivot.column_dimensions
            records += pivot.measures
            existing = [x for x in pivot.order]

            to_delete |= set(existing) - set(records)
            missing = set(records) - set(existing)
            for record in missing:
                to_save.append(
                    Order(pivot=pivot, element=record, sequence=sequence))

        Order.save(to_save)
        Order.delete(list(to_delete))


class RowDimension(sequence_ordered(), ModelSQL, ModelView):
    'Pivot Row Dimension'
    __name__ = 'babi.pivot.row_dimension'
    pivot = fields.Many2One('babi.pivot', 'Pivot', required=True,
        ondelete='CASCADE')
    field = fields.Many2One('babi.field', 'Field', required=True,
        ondelete='CASCADE', domain=[
                ('table', '=', Eval('table', -1)),
                ])
    table = fields.Function(fields.Many2One('babi.table', 'Table'),
        'on_change_with_table')

    @classmethod
    def __setup__(cls):
        super().__setup__()
        cls.__access__.add('pivot')

    def get_rec_name(self, name):
        return self.field.rec_name

    @fields.depends('pivot', '_parent_pivot.table')
    def on_change_with_table(self, name=None):
        if self.pivot and self.pivot.table:
            return self.pivot.table.id


class ColumnDimension(sequence_ordered(), ModelSQL, ModelView):
    'Pivot Column Dimension'
    __name__ = 'babi.pivot.column_dimension'
    pivot = fields.Many2One('babi.pivot', 'Pivot', required=True,
        ondelete='CASCADE')
    field = fields.Many2One('babi.field', 'Field', required=True,
        ondelete='CASCADE', domain=[
                ('table', '=', Eval('table', -1)),
                ])
    table = fields.Function(fields.Many2One('babi.table', 'Table'),
        'on_change_with_table')

    @classmethod
    def __setup__(cls):
        super().__setup__()
        cls.__access__.add('pivot')

    def get_rec_name(self, name):
        return self.field.rec_name

    @fields.depends('pivot', '_parent_pivot.table')
    def on_change_with_table(self, name=None):
        if self.pivot and self.pivot.table:
            return self.pivot.table.id


class Measure(sequence_ordered(), ModelSQL, ModelView):
    'Pivot Measure'
    __name__ = 'babi.pivot.measure'
    pivot = fields.Many2One('babi.pivot', 'Pivot', required=True,
        ondelete='CASCADE')
    field = fields.Many2One('babi.field', 'Field', required=True,
        ondelete='CASCADE', domain=[
                ('table', '=', Eval('table', -1)),
                ])
    table = fields.Function(fields.Many2One('babi.table', 'Table'),
        'on_change_with_table')
    aggregate = fields.Selection([
            ('sum', 'Sum'),
            ('avg', 'Average'),
            ('count', 'Count'),
            ('max', 'Max'),
            ('min', 'Min'),
            ], 'Aggregate')

    @staticmethod
    def default_aggregate():
        return 'sum'

    def get_rec_name(self, name):
        return self.field.rec_name

    @classmethod
    def __setup__(cls):
        super().__setup__()
        cls.__access__.add('pivot')

    @fields.depends('pivot', '_parent_pivot.table')
    def on_change_with_table(self, name=None):
        if self.pivot and self.pivot.table:
            return self.pivot.table.id


class Property(sequence_ordered(), ModelSQL, ModelView):
    'Pivot Property'
    __name__ = 'babi.pivot.property'
    pivot = fields.Many2One('babi.pivot', 'Pivot', required=True,
        ondelete='CASCADE')
    field = fields.Many2One('babi.field', 'Field', required=True,
        ondelete='CASCADE', domain=[
                ('table', '=', Eval('table', -1)),
                ])
    table = fields.Function(fields.Many2One('babi.table', 'Table'),
        'on_change_with_table')

    @classmethod
    def __setup__(cls):
        super().__setup__()
        cls.__access__.add('pivot')

    def get_rec_name(self, name):
        return self.field.rec_name

    @fields.depends('pivot', '_parent_pivot.table')
    def on_change_with_table(self, name=None):
        if self.pivot and self.pivot.table:
            return self.pivot.table.id


class Order(sequence_ordered(), ModelSQL, ModelView):
    'Pivot Order'
    __name__ = 'babi.pivot.order'
    pivot = fields.Many2One('babi.pivot', 'Pivot', required=True,
        ondelete='CASCADE')
    element = fields.Reference('Element', selection='get_elements',
        readonly=True)
    order = fields.Selection([
            ('asc', 'Ascending'),
            ('desc', 'Descending'),
            ], 'Order')
    table = fields.Function(fields.Many2One('babi.table', 'Table'),
        'on_change_with_table')

    @staticmethod
    def default_order():
        return 'asc'

    @classmethod
    def __setup__(cls):
        super().__setup__()
        cls.__access__.add('pivot')

    @fields.depends('pivot', '_parent_pivot.table')
    def on_change_with_table(self, name=None):
        if self.pivot and self.pivot.table:
            return self.pivot.table.id

    @classmethod
    def _get_elements(cls):
        return ['babi.pivot.row_dimension', 'babi.pivot.column_dimension',
            'babi.pivot.measure', 'babi.pivot.property']

    @classmethod
    def get_elements(cls):
        Model = Pool().get('ir.model')
        get_name = Model.get_name
        models = cls._get_elements()
        return [(None, '')] + [(m, get_name(m)) for m in models]


class PivotExcel(Report):
    'Pivot Excel Export'
    __name__ = 'babi.pivot.excel'

    @classmethod
    def __setup__(cls):
        super().__setup__()
        cls.__rpc__['execute'] = RPC(False)

    @classmethod
    def execute(cls, ids, data):
        pool = Pool()
        Pivot = pool.get('babi.pivot')
        Language = pool.get('ir.lang')

        if not ids:
            return
        cls.check_access()

        pivots = Pivot.browse(ids)

        language = Transaction().context.get('language', 'en')
        language, = Language.search([('code', '=', language)], limit=1)
        wb = Workbook()
        wb.remove(wb.active)
        for pivot in pivots:
            ws = wb.create_sheet(pivot.table.name)
            cube = pivot.get_cube()
            cube.column_expansions = Cube.EXPAND_ALL
            cube.row_expansions = Cube.EXPAND_ALL
            for row in cube.build():
                ws.append([x.formatted(language, excel=True) for x in row])

        if len(pivots) == 1:
            name = pivot.table.name
        else:
            name = gettext('babi.msg_pivots_filename')
        return ('xlsx', save_virtual_workbook(wb), False, name)


class OpenExecutionFiltered(StateView):

    def __init__(self):
        buttons = [
                Button('Cancel', 'end', 'tryton-cancel'),
                Button('Create', 'create_table', 'tryton-ok', True),
                ]
        super().__init__('babi.table', 0, buttons)

    def get_view(self, wizard, state_name):
        pool = Pool()
        Parameter = pool.get('babi.filter.parameter')
        Table = pool.get('babi.table')

        context = Transaction().context
        record = Table(context.get('active_id'))

        result = {}
        result['type'] = 'form'
        result['view_id'] = None
        result['model'] = 'babi.table'
        result['field_childs'] = None
        fields = {}
        parameter2report = {}

        # Check if record has 'filter' field to pass tests in version 7.2
        if record and hasattr(record, 'filter'):
            parameters = Parameter.search([('filter', '=', record.filter)])
        else:
            parameters = []

        parameters_to_remove = []
        for parameter in parameters:
            if not parameter.check_parameter_in_filter():
                parameters_to_remove.append(parameter)
        for parameter in parameters_to_remove:
            parameters.remove(parameter)

        encoder = PYSONEncoder()
        xml = '<form string="Create Parameterized Table">\n'
        xml += '<group id="filters" string="Filters" colspan="4">\n'
        for parameter in parameters:
            # The wizard breaks with non unicode data
            name = 'filter_parameter_%d' % parameter.id
            field_definition = {
                'loading': 'eager',
                'name': name,
                'string': parameter.name,
                'searchable': True,
                'create': True,
                'help': '',
                'context': {},
                'delete': True,
                'type': parameter.ttype,
                'select': False,
                'readonly': False,
                'required': True,
            }
            if parameter.ttype in ['many2one', 'many2many']:
                field_definition['relation'] = parameter.related_model.model
            if parameter2report:
                field_definition['states'] = {
                    'invisible': Not(In(Eval('report', 0),
                            parameter2report[parameter.filter.id])),
                    'required': In(Eval('report', 0),
                        parameter2report[parameter.filter.id]),
                    }
            else:
                field_definition['states'] = {}
            # Copied from Model.fields_get
            for attr in ('states', 'domain', 'context', 'digits', 'size',
                    'add_remove', 'format'):
                if attr in field_definition:
                    field_definition[attr] = encoder.encode(
                        field_definition[attr])

            if parameter.ttype == 'many2many':
                xml += '<field name="%s" colspan="4"/>\n' % (name)
            else:
                xml += '<label name="%s"/>\n' % (name)
                xml += '<field name="%s" colspan="3"/>\n' % (name)
            fields[name] = field_definition

        xml += '</group>\n'
        xml += '</form>\n'
        result['arch'] = xml
        result['fields'] = fields
        return result

    def get_defaults(self, wizard, state_name, fields):
        pool = Pool()
        Parameter = pool.get('babi.filter.parameter')
        context = Transaction().context
        model = context.get('active_model')

        defaults = {}
        parameters = Parameter.search([
                ('related_model.model', '=', model),
                ])
        for parameter in parameters:
            name = '%s_%d' % (parameter.name, parameter.id)
            defaults[name] = context.get('active_id')
        return defaults


class CustomDict(dict):

    def __getattr__(self, name):
        return {}

    def __setattr__(self, name, value):
        self[name] = value


class ParametrizeTable(Wizard):
    'Babi Parametrize Table'
    __name__ = 'babi.table.parametrize'

    start = OpenExecutionFiltered()
    create_table = StateAction('babi.act_babi_table_parametrized')

    def __getattribute__(self, name):
        if name == 'start':
            if not hasattr(self, 'filter_values'):
                self.filter_values = CustomDict()
            name = 'filter_values'
        return super().__getattribute__(name)

    def do_create_table(self, action):
        pool = Pool()
        Table = pool.get('babi.table')
        Parameter = pool.get('babi.filter.parameter')

        data = {}
        for key, value in self.filter_values.items():
            # Fields has id of the field appendend, so it must be removed.
            key = key.split('_')[-1]
            data[key] = value
        parameters = {x.id: x.name for x in Parameter.browse(data.keys())}
        params = {}
        for key, value in data.items():
            key = parameters.get(int(key))
            params[key] = value

        internal_name = 'parametrized_' + secrets.token_hex(8)
        table, = Table.copy([self.record], default={
                'internal_name': internal_name,
                'parameters': params,
                'crons': [],
                })
        with Transaction().set_context(queue_name=QUEUE_NAME):
            Table.__queue__._compute(table)
        action['res_id'] = table.id
        return action, {}
